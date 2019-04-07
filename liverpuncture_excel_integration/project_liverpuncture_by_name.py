#coding: utf-8
import openpyxl

# global parameter
patientADict = {}   # 保存源excel中的姓名
patientBDict = {}

# 将单元格内变量尽量以 float 类型返回
def get_value(text):
    try:
        return float(text)
    except ValueError:
        return text
    except TypeError:
        return text

# 将住院号写入字典并返回
def make_dict(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]
    new_dict = {}
    max_row = sheet.max_row

    for i in range(1, max_row+1):
        item = sheet.cell(i, 2).value
        new_dict[item] = i
    return new_dict

def backup_excel(src_name, backup_name):
    wb2 = openpyxl.Workbook()
    wb2.save(backup_name)
    print('新建成功')

    #读取数据
    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(backup_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]

    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数

    for m in range(1, max_row+1):
        for n in range(1, max_column+1):
            cell1 = sheet1.cell(m, n).value #获取data单元格数据
            sheet2.cell(m, n).value = cell1 #赋值到test单元格

    wb2.save(backup_name)   #保存数据
    wb1.close()
    wb2.close()

# src_name: 原始A表格
# ref_name: 参考表格B，A中没有B中有的信息会被添加到新excel中
# dst_name: 最终生成的新excel名称
def insert_excel(src_name, ref_name, dst_name):
    wb_d = openpyxl.Workbook()
    wb_d.save(dst_name)
    print('新建成功')

    #读取数据
    wb_s = openpyxl.load_workbook(src_name)
    wb_r = openpyxl.load_workbook(ref_name)
    wb_d = openpyxl.load_workbook(dst_name)
    sheet_s = wb_s[wb_s.sheetnames[0]]  # src sheet
    sheet_r = wb_r[wb_r.sheetnames[0]]  # ref sheet
    sheet_d = wb_d[wb_d.sheetnames[0]]  # dst sheet

    max_row = sheet_s.max_row       #最大行数
    max_col = sheet_s.max_column    #最大列数

    for m in range(1, max_row+1):
        patient_name = sheet_r.cell(m, 2).value
        if patient_name in patientADict:                # 如果能找到姓名
            for n in range(1, max_col+1):
                new_row = patientADict[patient_name]    # 根据 ref 表格中的名字，找到 src 表格中的行号
                cell_s = sheet_s.cell(new_row, n).value # 找出该行信息并填入新表中
                sheet_d.cell(m, n).value = cell_s
        else:
            for n in range(1, max_col+1):               # 找不到姓名就保留该行信息
                cell_r = sheet_r.cell(m, n).value
                sheet_d.cell(m, n).value = cell_r

    wb_d.save(dst_name)   #保存数据
    wb_s.close()
    wb_d.close()

if __name__ == "__main__":
    fileAName = "数据库.xlsx"
    fileBName = "需找到人名.xlsx"
    backAName = "结果.xlsx"
    patientADict = make_dict(fileAName)
    #patientBDict = make_dict(fileBName)
    insert_excel(fileAName, fileBName, backAName)
