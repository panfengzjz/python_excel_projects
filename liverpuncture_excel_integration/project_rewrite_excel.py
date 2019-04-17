#coding: utf-8
import openpyxl

# global parameter
patientList = {}   # 保存源excel中的姓名
type2_count = 0
type2_place = ""
patient_dict = {"0":0,
                "30":0,
                "60":0,
                "120":0,
                "180":0}
patient_str_dict = {"0":"",
                "30":"",
                "60":"",
                "120":"",
                "180":""}

# src_name: 原始表格名称
# ret_name: 最终生成的新excel名称
def rewrite_excel(src_name, ret_name):
    global type2_count
    global type2_place
    wb2 = openpyxl.Workbook()
    wb2.save(ret_name)
    print('新建成功')

    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(ret_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数
    
    i = 1
    patient_name = sheet1.cell(i, 3).value
    line_count = 1
    
    def init_count():
        global type2_count
        global type2_place
        global patient_dict
        global patient_str_dict
        type2_count = 0
        type2_place = ""
        patient_dict = {"0":0,
                        "30":0,
                        "60":0,
                        "120":0,
                        "180":0}
        patient_str_dict = {"0":"",
                        "30":"",
                        "60":"",
                        "120":"",
                        "180":""}
    init_count()

    while(i < max_row+1):
        cur_name = sheet1.cell(i, 3).value
        if (cur_name == patient_name):
            sample_type = sheet1.cell(i, 4).value
            cur_timepoint = str(sheet1.cell(i, 5).value)
            cur_amount = sheet1.cell(i, 6).value
            cur_place  = sheet1.cell(i, 7).value

            if (sample_type == 1):
                patient_dict[cur_timepoint] += cur_amount
                patient_str_dict[cur_timepoint] += "%s*%s, " %(cur_place, cur_amount)
            elif (sample_type == 2):
                type2_count += cur_amount
                type2_place += "%s*%s, " %(cur_place, cur_amount)
        else:
            def write_info():
                for col in range(1, 4):
                    sheet2.cell(line_count, col).value = sheet1.cell(i-1, col).value
                col4 = ""
                col5 = ""
                for key in patient_dict:
                    if (patient_dict[key] == 0):
                        continue
                    col4 += "%s'*%s, " %(key, patient_dict[key])
                    col5 += "%s'*%s(%s), " %(key, patient_dict[key], patient_str_dict[key])
                sheet2.cell(line_count, 4).value = col4
                sheet2.cell(line_count, 5).value = col5
                sheet2.cell(line_count, 6).value = type2_count
                sheet2.cell(line_count, 7).value = type2_place
            write_info()
            line_count += 1
            patient_name = cur_name
            init_count()
            continue
        i += 1
    write_info()

    wb2.save(ret_name)   #保存数据
    wb1.close()
    wb2.close()

if __name__ == "__main__":
    fileName = "计数sample.xlsx"
    reslName = "结果.xlsx"
    #patientList = make_dict(fileName)   #这次生成一个数组，方便pop
    #backup_excel(fileName, reslName)
    rewrite_excel(fileName, reslName)
