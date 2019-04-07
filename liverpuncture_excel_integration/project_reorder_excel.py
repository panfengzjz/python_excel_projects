#coding: utf-8
import openpyxl

# global parameter
patientList = []   # 保存源excel中的姓名

# 将住院号写入数组并返回
def make_list(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]
    new_list = []
    max_row = sheet.max_row

    for i in range(1, max_row+1):
        item = sheet.cell(i, 3).value  # 名字在第三列
        if item not in new_list:
            new_list.append(item)
    return new_list

# src_name: 原始表格名称
# ret_name: 最终生成的新excel名称
def reorder_excel(src_name, ret_name):
    wb2 = openpyxl.Workbook()
    wb2.save(ret_name)
    print('新建成功')

    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(ret_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数
    row_count = 1                   #保存当前 ret_name excel 中记录到的行数

    while(patientList):
        target_name = patientList.pop(0)        #当前循环需要找的名字
        for m in range(1, max_row+1):
            cur_flow_name = sheet1.cell(m, 3).value
            if (cur_flow_name != target_name):  #如果名字不同，则找下一行
                continue
            for n in range(1, max_column+1):    #如果相同则写入 ret_name 的下一行中
                cell1 = sheet1.cell(m, n).value #获取data单元格数据
                sheet2.cell(row_count, n).value = cell1 #赋值到test单元格
            row_count += 1

    wb2.save(ret_name)   #保存数据
    wb1.close()
    wb2.close()

if __name__ == "__main__":
    fileName = "标本整理project-sample.xlsx"
    reslName = "结果.xlsx"
    patientList = make_list(fileName)   #这次生成一个数组，方便pop
    reorder_excel(fileName, reslName)
