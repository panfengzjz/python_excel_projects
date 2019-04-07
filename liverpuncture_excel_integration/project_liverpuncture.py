#coding: utf-8
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import PatternFill

# global parameter
patientADict = {}
patientBDict = {}

# 将单元格内变量尽量以 float 类型返回
def get_value(text):
    try:
        return float(text)
    except ValueError:
        return text
    except TypeError:
        return text

# 将住院号写入字典并返回
def make_dict(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]
    new_dict = {}
    max_row = sheet.max_row

    for i in range(1, max_row+1):
        item = sheet.cell(i, 2).value
        new_dict[item] = i
    return new_dict

def read_excel(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def backup_excel(src_name, backup_name):
    wb2 = openpyxl.Workbook()
    wb2.save(backup_name)
    print('新建成功')

    #读取数据
    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(backup_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]

    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数

    for m in range(1, max_row+1):
        for n in range(1, max_column+1):
            cell1 = sheet1.cell(m, n).value #获取data单元格数据
            sheet2.cell(m, n).value = cell1 #赋值到test单元格

    wb2.save(backup_name)   #保存数据
    wb1.close()
    wb2.close()

# src_name: 原始A表格
# ref_name: 参考表格B，A中没有B中有的信息会被添加到新excel中
# dst_name: 最终生成的新excel名称
def insert_excel(src_name, ref_name, dst_name):
    pfill_empty = PatternFill(start_color ='FFFF00', 
                              end_color = 'FFFF00', 
                              fill_type = 'solid')  # 以黄色填充A表空白单元格
    pfill_cofil = PatternFill(start_color ='FF4500',
                              end_color = 'FF4500', 
                              fill_type = 'solid')  # 以红色填充AB表冲突单元格
    wb_d = openpyxl.Workbook()
    wb_d.save(dst_name)
    print('新建成功')

    #读取数据
    wb_s = openpyxl.load_workbook(src_name)
    wb_r = openpyxl.load_workbook(ref_name)
    wb_d = openpyxl.load_workbook(dst_name)
    sheet_s = wb_s[wb_s.sheetnames[0]]  # src sheet
    sheet_r = wb_r[wb_r.sheetnames[0]]  # ref sheet
    sheet_d = wb_d[wb_d.sheetnames[0]]  # dst sheet

    max_row = sheet_s.max_row       #最大行数
    max_col = sheet_s.max_column    #最大列数

    for m in range(1, max_row+1):
        for n in range(1, max_col+1):
            cell_s = get_value(sheet_s.cell(m, n).value)#获取src单元格数据
            patient_id = sheet_s.cell(m, 2).value       #找到病人的住院号
            try:
                new_row = patientBDict[patient_id]      #找到该病人在ref表格中的行号
                cell_r = sheet_r.cell(new_row, n).value #获取对应ref数据
            except KeyError:
                sheet_d.cell(m, n).value = cell_s       #若B表找不到该病人，填入A表内信息并找下一项
                continue
            cell_r = get_value(cell_r)

            if (cell_s == cell_r):
                sheet_d.cell(m, n).value = cell_r       #两表信息一样则填入信息、不做标记
            elif (cell_s == None) or (
                (type(cell_s) == type('')) and (cell_s.strip() == "")):
                sheet_d.cell(m, n).value = cell_r       #若A表为空，则填入B表信息，标为黄色
                sheet_d.cell(m, n).fill = pfill_empty
            else:
                sheet_d.cell(m, n).value = cell_s       #若两表信息冲突，填入A表信息，标为红色，并打印
                if (n == 1):
                    continue
                sheet_d.cell(m, n).fill = pfill_cofil
                print("(%d %d): (%s %s)" %(m, n, cell_s, cell_r))

    wb_d.save(dst_name)   #保存数据
    wb_s.close()
    wb_d.close()

if __name__ == "__main__":
    fileAName = "肝穿数据库-表A-0112sample.xlsx"
    fileBName = "合作数据库-表B-0112sample.xlsx"
    backAName = "表A-test.xlsx"
    #patientADict = make_dict(fileAName)
    patientBDict = make_dict(fileBName)
    insert_excel(fileAName, fileBName, backAName)
